import openpyxl
from openpyxl import Workbook
import matplotlib.pyplot as plt
import statistics

def save_table(data, excel_filename):
    # Convert the dictionary to an Excel file
    wb = Workbook()
    ws = wb.active
    ws.append(['Domain', 'URL', 'Price'])
    for domain, details in data.items():
        ws.append([domain, details['url'], details['value']])
    wb.save(excel_filename)

def save_table_image(excel_filename, image_filename):
    # Read the Excel file to create the table image
    wb = openpyxl.load_workbook(excel_filename)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = [[cell for cell in row] for row in ws.iter_rows(min_row=2, values_only=True)]
    
    fig, ax = plt.subplots(figsize=(12, len(rows) * 0.5))  # Adjust the figure size based on the number of rows
    ax.axis('tight')
    ax.axis('off')
    table = ax.table(cellText=rows, colLabels=headers, cellLoc='center', loc='center')
    
    # Adjust the font size
    table.auto_set_font_size(False)
    table.set_fontsize(10)
    table.scale(1.2, 1.2)  # Adjust the scale of the table
    
    # Save the table as an image with higher DPI
    plt.savefig(image_filename, bbox_inches='tight', pad_inches=0.1, dpi=300)

def filter_outliers(price_dict):
    prices = [float(value['value']) for value in price_dict.values() if 'value' in value]
    if len(prices) < 2:
        return price_dict

    mean = statistics.mean(prices)
    try:
        std_dev = statistics.stdev(prices)
        filtered_prices = [price for price in prices if abs(price - mean) <= 2 * std_dev]
    except statistics.StatisticsError:
        # If there are not enough data points to calculate stdev, return the original prices
        filtered_prices = prices

    filtered_price_dict = {key: value for key, value in price_dict.items() if float(value['value']) in filtered_prices}

    return filtered_price_dict

if __name__ == "__main__":
    # Example usage
    data = {
        'example.com': {'url': 'http://example.com', 'file_path': 'temp/example.com.html', 'correct_price': 100, 'value': 100},
        'example.org': {'url': 'http://example.org', 'file_path': 'temp/example.org.html', 'correct_price': 200, 'value': 200}
    }
    excel_filename = 'data.xlsx'
    image_filename = 'table_image.png'
    save_table(data, excel_filename)
    save_table_image(excel_filename, image_filename)